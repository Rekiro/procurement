import uuid
from datetime import datetime
from app.shared.timezone import IST

from fastapi import HTTPException, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.procurement.machinery_requests.models import (
    ProcMachineryRequest, ProcMachineryPurchaseOrder, ProcMachineryGrn,
)
from app.procurement.machinery_requests.schemas import (
    MachineryRequestCreate, MachineryFulfillRequest, MachineryRejectRequest,
    MachineryPoUpdateRequest, MachineryGrnCreate,
    MachineryInvoiceCreate, ApproveInvoiceRequest, RejectInvoiceRequest,
)
from app.procurement.vendors.models import ProcVendor
from app.procurement.invoices.models import ProcInvoice, ProcInvoicePoLink

MACHINERY_OPTIONS = [
    "Tower Crane", "Mobile Crane", "Excavator", "Bulldozer", "Forklift",
    "Generator", "Air Compressor", "Concrete Mixer", "Drill Machine",
    "Welding Machine", "Scaffolding", "Boom Lift", "Scissor Lift",
    "Dump Truck", "Concrete Pump", "Vibrator", "Compactor",
]


def get_machinery_options():
    return MACHINERY_OPTIONS


async def create_machinery_request(db: AsyncSession, data: MachineryRequestCreate,
                                    user_email: str):
    count = await db.scalar(select(func.count()).select_from(ProcMachineryRequest))
    requisition_id = f"MREQ-{(count or 0) + 1:05d}"

    req = ProcMachineryRequest(
        requisition_id=requisition_id,
        site_id=data.siteId,
        site_manager_email=user_email,
        justification=data.justification,
        items=[item.model_dump() for item in data.items],
    )
    db.add(req)
    await db.commit()
    return req


async def list_machinery_requests(db: AsyncSession, status_filter: str | None = None):
    q = select(ProcMachineryRequest).order_by(ProcMachineryRequest.created_at.desc())
    if status_filter:
        q = q.where(ProcMachineryRequest.status == status_filter)
    result = await db.execute(q)
    return list(result.scalars().all())


async def get_request_for_approval(db: AsyncSession, request_id: uuid.UUID):
    req = await db.get(ProcMachineryRequest, request_id)
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery request not found")
    return req


async def fulfill_machinery_request(db: AsyncSession, request_id: uuid.UUID,
                                     data: MachineryFulfillRequest):
    req = await db.get(ProcMachineryRequest, request_id)
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery request not found")
    if req.status != "PENDING_PH_APPROVAL":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Request is not in PENDING_PH_APPROVAL status")

    vendor = await db.get(ProcVendor, data.vendorId)
    if not vendor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vendor not found")

    now = datetime.now(IST)
    count = await db.scalar(select(func.count()).select_from(ProcMachineryPurchaseOrder))
    po_number = f"PO-MAC-{(count or 0) + 1:05d}"

    expected_delivery = None
    tat = None
    if data.expectedDeliveryDate:
        expected_delivery = datetime(
            data.expectedDeliveryDate.year,
            data.expectedDeliveryDate.month,
            data.expectedDeliveryDate.day,
            tzinfo=IST,
        )
        tat = (data.expectedDeliveryDate - now.date()).days

    po = ProcMachineryPurchaseOrder(
        po_number=po_number,
        machinery_request_id=req.id,
        vendor_id=vendor.id,
        site_id=req.site_id,
        region=data.region,
        po_date=now,
        expected_delivery_date=expected_delivery,
        tat=tat,
        tat_status="On Time" if tat and tat >= 0 else None,
        items=[item.model_dump() for item in data.items],
    )
    db.add(po)

    req.status = "PROCESSED"
    req.updated_at = now
    await db.commit()
    return po


async def reject_machinery_request(db: AsyncSession, request_id: uuid.UUID,
                                    data: MachineryRejectRequest):
    req = await db.get(ProcMachineryRequest, request_id)
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery request not found")
    if req.status != "PENDING_PH_APPROVAL":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Request is not in PENDING_PH_APPROVAL status")

    req.status = "REJECTED"
    req.rejection_reason = data.rejectionReason
    req.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(req)
    return req


async def list_machinery_pos(db: AsyncSession):
    result = await db.execute(
        select(ProcMachineryPurchaseOrder).order_by(ProcMachineryPurchaseOrder.created_at.desc())
    )
    return list(result.scalars().all())


async def get_machinery_po(db: AsyncSession, po_number: str):
    po = await db.scalar(
        select(ProcMachineryPurchaseOrder).where(
            ProcMachineryPurchaseOrder.po_number == po_number
        )
    )
    if not po:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery PO not found")
    return po


async def update_machinery_po(db: AsyncSession, po_number: str,
                               data: MachineryPoUpdateRequest):
    po = await get_machinery_po(db, po_number)

    if data.status is not None:
        po.status = data.status
        if data.status == "Delivered":
            po.date_of_delivery = datetime.now(IST)
    if data.deliveryType is not None:
        po.delivery_type = data.deliveryType
    if data.courierName is not None:
        po.courier_name = data.courierName
    if data.podNumber is not None:
        po.pod_number = data.podNumber

    po.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(po)
    return po


async def submit_machinery_grn(db: AsyncSession, po_number: str,
                                data: MachineryGrnCreate, user_email: str):
    po = await get_machinery_po(db, po_number)

    existing = await db.scalar(
        select(ProcMachineryGrn).where(ProcMachineryGrn.po_number == po_number)
    )
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="GRN already submitted for this PO")

    grn = ProcMachineryGrn(
        po_number=po_number,
        requestor_email=user_email,
        comments=data.comments,
        signed_dc_url=data.signedDcUrl,
        asset_condition_proof_url=data.assetConditionProofUrl,
        packaging_images=data.packagingImages,
    )
    db.add(grn)

    po.status = "Delivered"
    po.date_of_delivery = datetime.now(IST)
    po.signed_dc_url = data.signedDcUrl
    po.updated_at = datetime.now(IST)
    await db.commit()
    return grn


async def get_machinery_grn_evidence(db: AsyncSession, po_number: str):
    grn = await db.scalar(
        select(ProcMachineryGrn).where(ProcMachineryGrn.po_number == po_number)
    )
    if not grn:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="No GRN found for this PO")
    return grn


async def get_consolidated_items(db: AsyncSession, po_numbers: list[str]) -> list[dict]:
    items = []
    for po_number in po_numbers:
        po = await db.scalar(
            select(ProcMachineryPurchaseOrder).where(
                ProcMachineryPurchaseOrder.po_number == po_number
            )
        )
        if po and isinstance(po.items, list):
            for item in po.items:
                items.append({**item, "poNumber": po_number})
    return items


async def submit_machinery_invoice(db: AsyncSession, data: MachineryInvoiceCreate,
                                    vendor_email: str):
    vendor = await db.scalar(select(ProcVendor).where(ProcVendor.email == vendor_email))

    count = await db.scalar(select(func.count()).select_from(ProcInvoice))
    invoice_id = f"INV-{(count or 0) + 1:06d}"

    invoice = ProcInvoice(
        invoice_id=invoice_id,
        vendor_id=vendor.id if vendor else None,
        invoice_no=data.invoiceNo,
        invoice_type="machinery",
        state=data.state,
        bill_amount=data.billAmount,
        bill_url=data.billUrl,
    )
    db.add(invoice)
    await db.flush()

    for po_number in data.poNumbers:
        db.add(ProcInvoicePoLink(invoice_id=invoice.id, po_number=po_number))
        po = await db.scalar(
            select(ProcMachineryPurchaseOrder).where(
                ProcMachineryPurchaseOrder.po_number == po_number
            )
        )
        if po:
            po.status = "INVOICE_SUBMITTED"
            po.updated_at = datetime.now(IST)

    await db.commit()
    return invoice, data.poNumbers


async def list_machinery_invoices(db: AsyncSession):
    result = await db.execute(
        select(ProcInvoice)
        .where(ProcInvoice.invoice_type == "machinery")
        .order_by(ProcInvoice.submitted_at.desc())
    )
    return list(result.scalars().all())


async def approve_machinery_invoice(db: AsyncSession, data: ApproveInvoiceRequest,
                                     reviewed_by: str):
    invoice = await db.get(ProcInvoice, data.invoiceId)
    if not invoice or invoice.invoice_type != "machinery":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery invoice not found")
    if invoice.status != "Pending":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Invoice is not in Pending status")

    invoice.status = "Approved"
    invoice.reviewed_at = datetime.now(IST)
    invoice.reviewed_by = reviewed_by
    await db.commit()
    await db.refresh(invoice)
    return invoice


async def reject_machinery_invoice(db: AsyncSession, invoice_id: uuid.UUID,
                                    data: RejectInvoiceRequest, reviewed_by: str):
    invoice = await db.get(ProcInvoice, invoice_id)
    if not invoice or invoice.invoice_type != "machinery":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Machinery invoice not found")
    if invoice.status != "Pending":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Invoice is not in Pending status")

    invoice.status = "Rejected"
    invoice.rejection_reason = data.rejectionReason
    invoice.reviewed_at = datetime.now(IST)
    invoice.reviewed_by = reviewed_by
    await db.commit()
    await db.refresh(invoice)
    return invoice


async def get_vendor_name(db: AsyncSession, vendor_id: uuid.UUID | None) -> str | None:
    if not vendor_id:
        return None
    vendor = await db.get(ProcVendor, vendor_id)
    return vendor.company_name if vendor else None


async def export_machinery_pos_excel(db: AsyncSession, pos: list):
    from openpyxl import Workbook
    from app.shared.excel_utils import workbook_to_streaming_response, HEADER_FONT, HEADER_FILL, HEADER_BORDER
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Machinery Purchase Orders"

    headers = ["PO Number", "Site ID", "Region", "Status", "PO Date",
               "Expected Delivery", "Delivery Type", "Courier", "POD Number", "Date of Delivery"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)

    for po in pos:
        ws.append([
            po.po_number,
            po.site_id,
            po.region or "",
            po.status,
            po.po_date.strftime("%Y-%m-%d") if po.po_date else "",
            po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else "",
            po.delivery_type or "",
            po.courier_name or "",
            po.pod_number or "",
            po.date_of_delivery.strftime("%Y-%m-%d") if po.date_of_delivery else "",
        ])

    ws.freeze_panes = "A2"
    return workbook_to_streaming_response(wb, "machinery_purchase_orders.xlsx")

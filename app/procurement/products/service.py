import io
import math
from datetime import datetime, timezone, date

from fastapi import HTTPException, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.procurement.products.models import ProcProduct, ProcProductPriceChangeRequest
from app.procurement.vendors.models import ProcVendor
from app.procurement.products.schemas import (
    ProductCreate, ApproveProductRequest, RejectProductRequest,
    PriceChangeRequestCreate, ApprovePriceChangeRequest, RejectPriceChangeRequest,
)
from app.shared.excel_utils import create_template_workbook, workbook_to_streaming_response, parse_upload_to_rows


def _calculate_final_price(price: float, margin_pct: float | None, margin_amt: float | None) -> float:
    if margin_pct is not None:
        return round(price * (1 + margin_pct / 100), 2)
    if margin_amt is not None:
        return round(price + margin_amt, 2)
    return price


async def _get_vendor_by_code(db: AsyncSession, vendor_code: str) -> ProcVendor:
    vendor = await db.get(ProcVendor, vendor_code)
    if not vendor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Vendor {vendor_code} not found")
    if vendor.status != "ACTIVE":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Vendor account is not active")
    return vendor


async def _next_product_code(db: AsyncSession) -> str:
    count = await db.scalar(select(func.count()).select_from(ProcProduct))
    return f"PRD{(count or 0) + 1:07d}"


async def create_product(db: AsyncSession, data: ProductCreate) -> ProcProduct:
    vendor = await _get_vendor_by_code(db, data.vendorCode)
    product_code = await _next_product_code(db)

    product = ProcProduct(
        product_code=product_code,
        vendor_code=vendor.vendor_code,
        product_name=data.productName,
        category=data.category,
        subcategory=data.subcategory,
        price=data.price,
        hsn_code=data.hsnCode,
        is_tax_exempt=data.isTaxExempt,
        gst_rate=0.0 if data.isTaxExempt else data.gstRate,
        delivery_days=data.deliveryDays,
        delivery_cost=data.deliveryCost,
        uom=data.uom,
        description=data.description,
        final_price=data.price,   # no margin set yet
        status="Pending",
    )
    db.add(product)
    await db.commit()
    await db.refresh(product)
    return product


async def list_products(
    db: AsyncSession,
    status_filter: str | None = None,
    search: str | None = None,
    page: int = 1,
    limit: int = 10,
) -> tuple[list[tuple], dict]:
    """Returns (list of (ProcProduct, vendor_name) tuples, pagination dict)."""
    q = (
        select(ProcProduct, ProcVendor.company_name)
        .join(ProcVendor, ProcProduct.vendor_code == ProcVendor.vendor_code)
        .order_by(ProcProduct.created_at.desc())
    )
    if status_filter:
        q = q.where(ProcProduct.status == status_filter)
    if search:
        q = q.where(ProcProduct.product_name.ilike(f"%{search}%"))

    # Manual pagination (paginate() helper uses .scalars() which strips tuples)
    count_q = select(func.count()).select_from(q.subquery())
    total = await db.scalar(count_q) or 0

    if page < 1:
        page = 1
    if limit < 1:
        limit = 10
    offset = (page - 1) * limit
    result = await db.execute(q.offset(offset).limit(limit))
    rows = list(result.all())

    pagination = {
        "currentPage": page,
        "totalPages": math.ceil(total / limit) if limit else 1,
        "totalItems": total,
    }
    return rows, pagination


async def approve_products(db: AsyncSession, data: ApproveProductRequest, reviewed_by: str) -> list[ProcProduct]:
    approved = []
    for product_code in data.productIds:
        product = await db.get(ProcProduct, product_code)
        if not product:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Product {product_code} not found")
        if product.status != "Pending":
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Product {product_code} is already {product.status}")
        product.status = "Approved"
        product.updated_at = datetime.now(timezone.utc)
        approved.append(product)

    await db.commit()
    for p in approved:
        await db.refresh(p)
    return approved


async def reject_product(db: AsyncSession, product_code: str, data: RejectProductRequest, reviewed_by: str) -> ProcProduct:
    product = await db.get(ProcProduct, product_code)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")

    product.status = "Rejected"
    product.rejection_reason = data.reason
    product.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(product)
    return product


async def get_product_catalog(db: AsyncSession) -> list[tuple]:
    result = await db.execute(
        select(ProcProduct, ProcVendor.company_name)
        .join(ProcVendor, ProcProduct.vendor_code == ProcVendor.vendor_code)
        .where(ProcProduct.status == "Approved")
        .order_by(ProcProduct.product_name)
    )
    return list(result.all())


async def list_vendor_products(
    db: AsyncSession,
    vendor_code: str,
    page: int = 1,
    limit: int = 10,
) -> tuple[list[ProcProduct], dict]:
    q = (
        select(ProcProduct)
        .where(ProcProduct.vendor_code == vendor_code)
        .order_by(ProcProduct.created_at.desc())
    )
    count_q = select(func.count()).select_from(q.subquery())
    total = await db.scalar(count_q) or 0

    offset = (page - 1) * limit
    result = await db.execute(q.offset(offset).limit(limit))
    products = list(result.scalars().all())

    pagination = {
        "currentPage": page,
        "totalPages": math.ceil(total / limit) if limit else 1,
        "totalItems": total,
    }
    return products, pagination


async def delete_product(db: AsyncSession, product_code: str, vendor_code: str) -> str:
    from app.procurement.purchase_orders.models import ProcPoItem, ProcPurchaseOrder
    from app.procurement.indents.models import ProcIndent, ProcIndentItem

    product = await db.get(ProcProduct, product_code)
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    if product.vendor_code != vendor_code:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You don't own this product")

    # Block if product is in a live indent (pre-PO stage)
    active_indent_statuses = {"PENDING_PH_APPROVAL", "PENDING_RM_APPROVAL", "PH_APPROVED"}
    active_indent = await db.scalar(
        select(ProcIndent.tracking_no)
        .join(ProcIndentItem, ProcIndentItem.indent_id == ProcIndent.id)
        .where(
            ProcIndentItem.product_code == product_code,
            ProcIndent.status.in_(active_indent_statuses),
        )
        .limit(1)
    )
    if active_indent:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete product: it is part of active indent {active_indent}.",
        )

    # Block if product is in an active unfulfilled Purchase Order
    active_po_statuses = {"Pending", "Confirmed", "Processing", "Shipped"}
    active_po = await db.scalar(
        select(ProcPurchaseOrder.po_number)
        .join(ProcPoItem, ProcPoItem.po_id == ProcPurchaseOrder.id)
        .where(
            ProcPoItem.product_code == product_code,
            ProcPurchaseOrder.status.in_(active_po_statuses),
        )
        .limit(1)
    )
    if active_po:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete product: it is referenced by active Purchase Order {active_po}.",
        )

    # Cascade delete price change requests (FK constraint)
    pcrs = await db.execute(
        select(ProcProductPriceChangeRequest).where(
            ProcProductPriceChangeRequest.product_code == product_code
        )
    )
    for pcr in pcrs.scalars().all():
        await db.delete(pcr)

    await db.delete(product)
    await db.commit()
    return product_code


async def create_price_change_request(db: AsyncSession, data: PriceChangeRequestCreate) -> ProcProductPriceChangeRequest:
    vendor = await _get_vendor_by_code(db, data.vendorCode)

    if data.wefDate <= date.today():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="wefDate must be a future date")

    product = await db.get(ProcProduct, data.productId)
    if not product or product.vendor_code != vendor.vendor_code:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found or not owned by vendor")

    existing = await db.scalar(
        select(ProcProductPriceChangeRequest).where(
            ProcProductPriceChangeRequest.product_code == data.productId,
            ProcProductPriceChangeRequest.status == "Pending",
        )
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"A pending price change request already exists for product {data.productId} (approvalId: {existing.id}). Approve or reject it before submitting a new one.",
        )

    approval_id = f"PROD-EDIT-{int(datetime.now(timezone.utc).timestamp() * 1000)}"
    pcr = ProcProductPriceChangeRequest(
        id=approval_id,
        product_code=data.productId,
        vendor_code=vendor.vendor_code,
        new_price=data.newPrice,
        wef_date=data.wefDate,
        status="Pending",
    )
    db.add(pcr)
    await db.commit()
    await db.refresh(pcr)
    return pcr


async def list_price_change_requests(
    db: AsyncSession,
    status_filter: str | None = None,
    search: str | None = None,
    page: int = 1,
    limit: int = 10,
) -> tuple[list[tuple], dict]:
    """Returns (list of (PCR, product_name, vendor_name, original_price) tuples, pagination dict)."""
    q = (
        select(
            ProcProductPriceChangeRequest,
            ProcProduct.product_name,
            ProcVendor.company_name,
            ProcProduct.price,
        )
        .join(ProcProduct, ProcProductPriceChangeRequest.product_code == ProcProduct.product_code)
        .join(ProcVendor, ProcProductPriceChangeRequest.vendor_code == ProcVendor.vendor_code)
        .order_by(ProcProductPriceChangeRequest.created_at.desc())
    )
    if status_filter:
        q = q.where(ProcProductPriceChangeRequest.status == status_filter)
    if search:
        term = f"%{search}%"
        q = q.where(
            ProcProduct.product_name.ilike(term)
            | ProcProductPriceChangeRequest.product_code.ilike(term)
        )

    count_q = select(func.count()).select_from(q.subquery())
    total = await db.scalar(count_q) or 0

    offset = (page - 1) * limit
    result = await db.execute(q.offset(offset).limit(limit))
    rows = list(result.all())

    pagination = {
        "currentPage": page,
        "totalPages": math.ceil(total / limit) if limit else 1,
        "totalItems": total,
    }
    return rows, pagination


async def approve_price_change_requests(
    db: AsyncSession, data: ApprovePriceChangeRequest, reviewed_by: str,
) -> list[ProcProductPriceChangeRequest]:
    result = await db.execute(
        select(ProcProductPriceChangeRequest).where(
            ProcProductPriceChangeRequest.id.in_(data.approvalIds)
        )
    )
    pcrs = list(result.scalars().all())

    not_found = set(data.approvalIds) - {pcr.id for pcr in pcrs}
    if not_found:
        raise HTTPException(status_code=404, detail=f"Price change requests not found: {sorted(not_found)}")

    not_pending = [pcr.id for pcr in pcrs if pcr.status != "Pending"]
    if not_pending:
        raise HTTPException(status_code=409, detail=f"Requests not in Pending status: {not_pending}")

    now = datetime.now(timezone.utc)
    for pcr in pcrs:
        # NOTE: spec says schedule price update for wefDate; no scheduler infra yet — applying immediately
        product = await db.get(ProcProduct, pcr.product_code)
        if product:
            product.price = pcr.new_price
            product.final_price = _calculate_final_price(
                float(pcr.new_price),
                float(product.margin_percentage) if product.margin_percentage else None,
                float(product.direct_margin_amount) if product.direct_margin_amount else None,
            )
            product.updated_at = now
        pcr.status = "Approved"
        pcr.reviewed_at = now
        pcr.reviewed_by = reviewed_by

    await db.commit()
    for pcr in pcrs:
        await db.refresh(pcr)
    return pcrs


async def reject_price_change_request(
    db: AsyncSession, approval_id: str, data: RejectPriceChangeRequest, reviewed_by: str,
) -> ProcProductPriceChangeRequest:
    pcr = await db.get(ProcProductPriceChangeRequest, approval_id)
    if not pcr:
        raise HTTPException(status_code=404, detail="Price change request not found")
    if pcr.status != "Pending":
        raise HTTPException(status_code=409, detail=f"Request is already {pcr.status}")

    pcr.status = "Rejected"
    pcr.rejection_reason = data.reason
    pcr.reviewed_at = datetime.now(timezone.utc)
    pcr.reviewed_by = reviewed_by
    await db.commit()
    await db.refresh(pcr)
    return pcr


async def get_margins(
    db: AsyncSession,
    page: int = 1,
    limit: int = 10,
) -> tuple[list[ProcProduct], dict]:
    from datetime import timedelta
    q = (
        select(ProcProduct)
        .where(ProcProduct.status == "Approved")
        .order_by(ProcProduct.product_name)
    )
    count_q = select(func.count()).select_from(q.subquery())
    total = await db.scalar(count_q) or 0

    offset = (page - 1) * limit
    result = await db.execute(q.offset(offset).limit(limit))
    products = list(result.scalars().all())

    pagination = {
        "currentPage": page,
        "totalPages": math.ceil(total / limit) if limit else 1,
        "totalItems": total,
    }
    return products, pagination


async def get_margin_template(db: AsyncSession):
    from datetime import date as date_type, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Protection

    today = date_type.today()
    result = await db.execute(
        select(ProcProduct).where(ProcProduct.status == "Approved").order_by(ProcProduct.product_name)
    )
    products = list(result.scalars().all())

    headers = ["Product ID", "Product Name", "Category", "Price", "Margin (%)", "Direct Margin (₹)"]
    wb = create_template_workbook(headers)
    ws = wb.active

    for row_idx, p in enumerate(products, start=2):
        delivery_date = today + timedelta(days=p.delivery_days)
        ws.cell(row=row_idx, column=1, value=p.product_code)
        ws.cell(row=row_idx, column=2, value=p.product_name)
        ws.cell(row=row_idx, column=3, value=p.category)
        ws.cell(row=row_idx, column=4, value=float(p.price))
        ws.cell(row=row_idx, column=5, value=float(p.margin_percentage) if p.margin_percentage is not None else None)
        ws.cell(row=row_idx, column=6, value=float(p.direct_margin_amount) if p.direct_margin_amount is not None else None)

    filename = f"Product_Margins_Template_{today.strftime('%Y-%m-%d')}.xlsx"
    return workbook_to_streaming_response(wb, filename)


async def get_product_bulk_upload_template(db: AsyncSession):
    # Query distinct categories and subcategories from approved products
    cat_result = await db.execute(
        select(ProcProduct.category).where(ProcProduct.category.isnot(None)).distinct()
    )
    categories = sorted([r[0] for r in cat_result.all() if r[0]])

    subcat_result = await db.execute(
        select(ProcProduct.subcategory).where(ProcProduct.subcategory.isnot(None)).distinct()
    )
    subcategories = sorted([r[0] for r in subcat_result.all() if r[0]])

    headers = [
        "Product Name", "Category", "Sub Category", "Price", "HSN Code",
        "Is Tax Exempt", "GST Rate (%)", "UOM", "Number of Delivery Days", "Cost of Delivery", "Description",
    ]
    dropdowns = {
        "Category": categories or ["General"],
        "Sub Category": subcategories or ["General"],
        "Is Tax Exempt": ["Yes", "No"],
        "UOM": ["PCS", "KG", "LTR", "BOX", "MTR", "SET"],
    }
    wb = create_template_workbook(headers, dropdowns)
    return workbook_to_streaming_response(wb, "Product_Upload_Template.xlsx")


async def bulk_upload_margins(db: AsyncSession, file_bytes: bytes, filename: str, reviewed_by: str) -> dict:
    from datetime import date as date_type, timedelta

    rows = parse_upload_to_rows(file_bytes, filename)
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows.")

    # Verify required headers are present in the file
    required_headers = {"Product ID", "Margin (%)", "Direct Margin (₹)"}
    actual_headers = set(rows[0].keys())
    missing_headers = required_headers - actual_headers
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required column(s): {', '.join(sorted(missing_headers))}. "
                   f"Please use the template downloaded from the system.",
        )

    def _parse_num(val) -> float | None:
        if val in (None, "", "None"):
            return None
        return float(val)

    # --- Phase 1: validate ALL rows (including product existence) ---
    errors = []
    valid_rows: list[tuple[int, str, ProcProduct, float | None, float | None]] = []

    for i, row in enumerate(rows, start=2):
        product_code = str(row.get("Product ID", "") or "").strip()
        product_name = str(row.get("Product Name", "") or "").strip()

        if not product_code:
            errors.append({"rowIndex": i, "productName": product_name or "(blank)", "error": "Product ID is required."})
            continue

        # Find product in DB — part of per-row validation
        product = await db.scalar(select(ProcProduct).where(ProcProduct.product_code == product_code))
        if not product:
            errors.append({"rowIndex": i, "productName": product_name or product_code, "error": f"Product '{product_code}' not found in the system."})
            continue

        # Check both margin columns
        margin_pct_raw = row.get("Margin (%)")
        margin_amt_raw = row.get("Direct Margin (₹)")

        try:
            margin_pct = _parse_num(margin_pct_raw)
        except (ValueError, TypeError):
            errors.append({"rowIndex": i, "productName": product.product_name, "error": "Margin (%) must be a valid number."})
            continue

        try:
            margin_amt = _parse_num(margin_amt_raw)
        except (ValueError, TypeError):
            errors.append({"rowIndex": i, "productName": product.product_name, "error": "Direct Margin (₹) must be a valid number."})
            continue

        if margin_pct is not None and margin_amt is not None:
            errors.append({"rowIndex": i, "productName": product.product_name, "error": "Only one of Margin (%) or Direct Margin (₹) can be set, not both."})
            continue

        if margin_pct is None and margin_amt is None:
            continue  # skip rows with no margin values

        valid_rows.append((i, product_code, product, margin_pct, margin_amt))

    if errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"File validation failed. Found {len(errors)} error(s). Please correct the file and re-upload.",
                "errors": errors,
            },
        )

    if not valid_rows:
        raise HTTPException(
            status_code=400,
            detail="No margin data found in the file. Please fill in at least one Margin (%) or Direct Margin (₹) value before uploading.",
        )

    # --- Phase 2: apply all updates (all rows already validated) ---
    today = date_type.today()
    updated_products = []
    now = datetime.now(timezone.utc)

    for _, product_code, product, margin_pct, margin_amt in valid_rows:

        product.margin_percentage = margin_pct
        product.direct_margin_amount = margin_amt
        product.final_price = _calculate_final_price(float(product.price), margin_pct, margin_amt)
        product.updated_at = now

        final_with_margin = float(product.final_price) if (margin_pct is not None or margin_amt is not None) else None
        updated_products.append({
            "id": product.product_code,
            "title": product.product_name,
            "category": product.category,
            "price": float(product.price),
            "deliveryDate": (today + timedelta(days=product.delivery_days)).isoformat(),
            "marginPercentage": margin_pct,
            "directMarginAmount": margin_amt,
            "finalPriceWithMargin": final_with_margin,
        })

    await db.commit()
    return {
        "message": f"Margins for {len(updated_products)} product(s) have been successfully updated.",
        "updatedProducts": updated_products,
    }


async def bulk_upload_products(db: AsyncSession, file_bytes: bytes, filename: str, vendor_code: str) -> dict:
    from fastapi.responses import JSONResponse

    vendor = await _get_vendor_by_code(db, vendor_code)

    # Pre-fetch valid categories and UOM list for validation
    cat_result = await db.execute(
        select(ProcProduct.category).where(ProcProduct.category.isnot(None)).distinct()
    )
    valid_categories = {r[0] for r in cat_result.all() if r[0]}
    valid_uom = {"PCS", "KG", "LTR", "BOX", "MTR", "SET"}

    rows = parse_upload_to_rows(file_bytes, filename)
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows.")

    # --- Phase 1: validate ALL rows, collect errors ---
    errors = []
    for i, row in enumerate(rows, start=2):
        product_name = str(row.get("Product Name", "") or "").strip()
        price_raw = row.get("Price")
        hsn_code = str(row.get("HSN Code", "") or "").strip()
        category = str(row.get("Category", "") or "").strip()
        uom = str(row.get("UOM", "") or "").strip()
        delivery_days_raw = row.get("Number of Delivery Days")

        if not product_name:
            errors.append({"rowIndex": i, "productName": product_name or "(blank)", "error": "Product Name is required."})
            continue

        try:
            price = float(price_raw)
            if price <= 0:
                errors.append({"rowIndex": i, "productName": product_name, "error": "Price must be a positive number."})
        except (TypeError, ValueError):
            errors.append({"rowIndex": i, "productName": product_name, "error": "Price must be a valid number."})

        if not hsn_code or not hsn_code.isdigit() or len(hsn_code) != 4:
            errors.append({"rowIndex": i, "productName": product_name, "error": "HSN Code must be exactly 4 digits."})

        if not category:
            errors.append({"rowIndex": i, "productName": product_name, "error": "Category is required."})
        elif valid_categories and category not in valid_categories:
            errors.append({"rowIndex": i, "productName": product_name, "error": f"Category '{category}' is not a valid option."})

        if uom not in valid_uom:
            errors.append({"rowIndex": i, "productName": product_name, "error": f"UOM must be one of: {', '.join(sorted(valid_uom))}."})

        is_tax_exempt_raw = str(row.get("Is Tax Exempt", "") or "").strip().lower()
        is_tax_exempt = is_tax_exempt_raw in ("yes", "true", "1")
        gst_rate_raw = row.get("GST Rate (%)")
        try:
            gst_rate = float(gst_rate_raw or 0)
        except (TypeError, ValueError):
            gst_rate = 0
        if is_tax_exempt and gst_rate > 0:
            errors.append({"rowIndex": i, "productName": product_name, "error": "GST Rate must be 0 when Is Tax Exempt is 'Yes'."})

        try:
            days = int(delivery_days_raw)
            if days < 1:
                errors.append({"rowIndex": i, "productName": product_name, "error": "Number of Delivery Days must be at least 1."})
        except (TypeError, ValueError):
            errors.append({"rowIndex": i, "productName": product_name, "error": "Number of Delivery Days must be a valid integer."})

    if errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": f"File validation failed. Found {len(errors)} error(s). Please correct the file and re-upload.",
                "errors": errors,
            },
        )

    # --- Phase 2: all rows valid — insert ---
    for row in rows:
        product_code = await _next_product_code(db)
        price = float(row.get("Price", 0))
        product = ProcProduct(
            product_code=product_code,
            vendor_code=vendor.vendor_code,
            product_name=str(row.get("Product Name", "")).strip(),
            category=str(row.get("Category", "")).strip(),
            subcategory=str(row.get("Sub Category", "") or "").strip(),
            price=price,
            hsn_code=str(row.get("HSN Code", "")).strip(),
            is_tax_exempt=str(row.get("Is Tax Exempt", "") or "").strip().lower() in ("yes", "true", "1"),
            gst_rate=float(row.get("GST Rate (%)", 0) or 0),
            delivery_days=int(row.get("Number of Delivery Days", 1)),
            delivery_cost=float(row.get("Cost of Delivery", 0) or 0),
            uom=str(row.get("UOM", "PCS")).strip(),
            description=str(row.get("Description", "") or "").strip() or None,
            final_price=price,
            status="Pending",
        )
        db.add(product)
        await db.flush()

    await db.commit()
    return {"message": f"File processed successfully. {len(rows)} product(s) have been submitted for approval."}

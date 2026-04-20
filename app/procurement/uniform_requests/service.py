import uuid
from datetime import datetime
from app.shared.timezone import IST

from fastapi import HTTPException, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.procurement.uniform_requests.models import ProcUniformRequest, ProcUniformPurchaseOrder
from app.procurement.uniform_requests.schemas import (
    UniformRequestCreate, UniformFulfillRequest, UniformRejectRequest,
    UniformPoUpdateRequest, UniformGrnCreate,
    UniformInvoiceCreate, ApproveInvoiceRequest, RejectInvoiceRequest,
)
from app.procurement.vendors.models import ProcVendor
from app.procurement.invoices.models import ProcInvoice, ProcInvoicePoLink

UNIFORM_CONFIGURATION = {
    "itemTypes": [
        {"name": "Shirt", "sizes": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]},
        {"name": "Trouser", "sizes": ["28", "30", "32", "34", "36", "38", "40"]},
        {"name": "T-Shirt", "sizes": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]},
        {"name": "Jacket", "sizes": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"]},
        {"name": "Safety Shoes", "sizes": ["6", "7", "8", "9", "10", "11", "12"]},
        {"name": "Safety Helmet", "sizes": ["Standard"]},
        {"name": "Safety Vest", "sizes": ["S", "M", "L", "XL", "XXL"]},
        {"name": "Gloves", "sizes": ["S", "M", "L", "XL"]},
        {"name": "Cap", "sizes": ["Standard"]},
        {"name": "Belt", "sizes": ["28", "30", "32", "34", "36", "38", "40"]},
    ],
    "issueTypes": ["new", "replacement", "backfill"],
}


def get_uniform_configuration():
    return UNIFORM_CONFIGURATION


async def search_employees(query: str) -> list[dict]:
    # Mock employee search — in production this would query the commercial HR tables
    mock_employees = [
        {"employeeCode": "EMP-001", "employeeName": "Rahul Sharma", "designation": "Security Guard", "site": "SITE-001"},
        {"employeeCode": "EMP-002", "employeeName": "Priya Singh", "designation": "Supervisor", "site": "SITE-001"},
        {"employeeCode": "EMP-003", "employeeName": "Amit Kumar", "designation": "Guard", "site": "SITE-002"},
        {"employeeCode": "EMP-004", "employeeName": "Suresh Patel", "designation": "Manager", "site": "SITE-002"},
    ]
    if not query:
        return mock_employees
    q = query.lower()
    return [e for e in mock_employees if q in e["employeeName"].lower() or q in e["employeeCode"].lower()]


async def create_uniform_request(db: AsyncSession, data: UniformRequestCreate, user_email: str):
    count = await db.scalar(select(func.count()).select_from(ProcUniformRequest))
    request_id = f"UNF-{(count or 0) + 1:05d}"

    req = ProcUniformRequest(
        request_id=request_id,
        employee_code=data.employeeCode,
        employee_name=data.employeeName,
        designation=data.designation,
        site=data.site,
        client=data.client,
        issue_type=data.issueType,
        replacing_employee_code=data.replacingEmployeeCode,
        justification=data.justification,
        is_early_replacement=data.isEarlyReplacement,
        items=[item.model_dump() for item in data.items],
    )
    db.add(req)
    await db.commit()
    return req


async def list_uniform_requests(db: AsyncSession, status_filter: str | None = None):
    q = select(ProcUniformRequest).order_by(ProcUniformRequest.created_at.desc())
    if status_filter:
        q = q.where(ProcUniformRequest.status == status_filter)
    result = await db.execute(q)
    return list(result.scalars().all())


async def get_employee_uniform_history(db: AsyncSession, employee_code: str):
    result = await db.execute(
        select(ProcUniformRequest)
        .where(ProcUniformRequest.employee_code == employee_code)
        .order_by(ProcUniformRequest.created_at.desc())
    )
    return list(result.scalars().all())


async def fulfill_uniform_request(db: AsyncSession, request_id: uuid.UUID,
                                   data: UniformFulfillRequest):
    req = await db.get(ProcUniformRequest, request_id)
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Uniform request not found")
    if req.status != "PENDING_PH_APPROVAL":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Request is not in PENDING_PH_APPROVAL status")

    vendor = await db.get(ProcVendor, data.vendorId)
    if not vendor:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vendor not found")

    now = datetime.now(IST)
    count = await db.scalar(select(func.count()).select_from(ProcUniformPurchaseOrder))
    po_number = f"PO-UNF-{(count or 0) + 1:05d}"

    expected_delivery = None
    tat = None
    if data.expectedDeliveryDate:
        expected_delivery = datetime(
            data.expectedDeliveryDate.year,
            data.expectedDeliveryDate.month,
            data.expectedDeliveryDate.day,
            tzinfo=IST,
        )
        tat = (data.expectedDeliveryDate - now.date()).days

    po = ProcUniformPurchaseOrder(
        po_number=po_number,
        uniform_request_id=req.id,
        vendor_id=vendor.id,
        employee_name=req.employee_name,
        employee_code=req.employee_code,
        site_name=req.site,
        region=data.region,
        po_date=now,
        expected_delivery_date=expected_delivery,
        tat=tat,
        tat_status="On Time" if tat and tat >= 0 else None,
        items=[item.model_dump() for item in data.items],
    )
    db.add(po)

    req.status = "PROCESSED"
    req.updated_at = now
    await db.commit()
    return po


async def reject_uniform_request(db: AsyncSession, request_id: uuid.UUID,
                                  data: UniformRejectRequest):
    req = await db.get(ProcUniformRequest, request_id)
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Uniform request not found")
    if req.status != "PENDING_PH_APPROVAL":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Request is not in PENDING_PH_APPROVAL status")

    req.status = "REJECTED"
    req.rejection_reason = data.rejectionReason
    req.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(req)
    return req


async def list_uniform_pos(db: AsyncSession):
    result = await db.execute(
        select(ProcUniformPurchaseOrder).order_by(ProcUniformPurchaseOrder.created_at.desc())
    )
    return list(result.scalars().all())


async def get_uniform_po(db: AsyncSession, po_number: str):
    po = await db.scalar(
        select(ProcUniformPurchaseOrder).where(
            ProcUniformPurchaseOrder.po_number == po_number
        )
    )
    if not po:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND,
                            detail="Uniform PO not found")
    return po


async def update_uniform_po(db: AsyncSession, po_number: str, data: UniformPoUpdateRequest):
    po = await get_uniform_po(db, po_number)

    if data.status is not None:
        po.status = data.status
        if data.status == "Delivered":
            po.date_of_delivery = datetime.now(IST)
    if data.deliveryType is not None:
        po.delivery_type = data.deliveryType
    if data.courierName is not None:
        po.courier_name = data.courierName
    if data.podNumber is not None:
        po.pod_number = data.podNumber

    po.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(po)
    return po


async def submit_uniform_grn(db: AsyncSession, po_number: str, data: UniformGrnCreate,
                              user_email: str):
    po = await get_uniform_po(db, po_number)
    if po.status == "Delivered":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="GRN already submitted for this PO")

    po.status = "Delivered"
    po.date_of_delivery = datetime.now(IST)
    po.signed_dc_url = data.signedDcUrl
    po.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(po)
    return po


async def get_consolidated_items(db: AsyncSession, po_numbers: list[str]) -> list[dict]:
    items = []
    for po_number in po_numbers:
        po = await db.scalar(
            select(ProcUniformPurchaseOrder).where(
                ProcUniformPurchaseOrder.po_number == po_number
            )
        )
        if po and isinstance(po.items, list):
            for item in po.items:
                items.append({**item, "poNumber": po_number})
    return items


async def submit_uniform_invoice(db: AsyncSession, data: UniformInvoiceCreate,
                                  vendor_email: str):
    vendor = await db.scalar(select(ProcVendor).where(ProcVendor.email == vendor_email))

    count = await db.scalar(select(func.count()).select_from(ProcInvoice))
    invoice_id = f"INV-{(count or 0) + 1:06d}"

    invoice = ProcInvoice(
        invoice_id=invoice_id,
        vendor_id=vendor.id if vendor else None,
        invoice_no=data.invoiceNo,
        invoice_type="uniform",
        state=data.state,
        bill_amount=data.billAmount,
        bill_url=data.billUrl,
    )
    db.add(invoice)
    await db.flush()

    for po_number in data.poNumbers:
        db.add(ProcInvoicePoLink(invoice_id=invoice.id, po_number=po_number))
        po = await db.scalar(
            select(ProcUniformPurchaseOrder).where(
                ProcUniformPurchaseOrder.po_number == po_number
            )
        )
        if po:
            po.status = "INVOICE_SUBMITTED"
            po.updated_at = datetime.now(IST)

    await db.commit()
    return invoice, data.poNumbers


async def list_uniform_invoices(db: AsyncSession):
    result = await db.execute(
        select(ProcInvoice)
        .where(ProcInvoice.invoice_type == "uniform")
        .order_by(ProcInvoice.submitted_at.desc())
    )
    return list(result.scalars().all())


async def get_vendor_name(db: AsyncSession, vendor_id: uuid.UUID | None) -> str | None:
    if not vendor_id:
        return None
    vendor = await db.get(ProcVendor, vendor_id)
    return vendor.company_name if vendor else None


async def export_uniform_pos_excel(db: AsyncSession, pos: list):
    from openpyxl import Workbook
    from app.shared.excel_utils import workbook_to_streaming_response, HEADER_FONT, HEADER_FILL, HEADER_BORDER
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Uniform Purchase Orders"

    headers = ["PO Number", "Employee Name", "Employee Code", "Site", "Region",
               "Status", "PO Date", "Expected Delivery", "Delivery Type",
               "Courier", "POD Number", "Date of Delivery"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 16)

    for po in pos:
        ws.append([
            po.po_number,
            po.employee_name,
            po.employee_code,
            po.site_name,
            po.region or "",
            po.status,
            po.po_date.strftime("%Y-%m-%d") if po.po_date else "",
            po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else "",
            po.delivery_type or "",
            po.courier_name or "",
            po.pod_number or "",
            po.date_of_delivery.strftime("%Y-%m-%d") if po.date_of_delivery else "",
        ])

    ws.freeze_panes = "A2"
    return workbook_to_streaming_response(wb, "uniform_purchase_orders.xlsx")

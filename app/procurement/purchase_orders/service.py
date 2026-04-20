import math
import uuid
from datetime import datetime
from app.shared.timezone import IST

from fastapi import HTTPException, status
from sqlalchemy import select, func, or_, cast, String
from sqlalchemy.ext.asyncio import AsyncSession

from app.procurement.purchase_orders.models import (
    ProcPurchaseOrder, ProcPoItem, ProcGrn, ProcGrnItem, ProcGrnPhoto,
)
from app.procurement.purchase_orders.schemas import PoUpdateData, GrnData
from app.procurement.indents.models import ProcIndent
from app.procurement.vendors.models import ProcVendor
from app.procurement.sites.models import Site


async def get_po_with_items(db: AsyncSession, po_number: str):
    po = await db.scalar(
        select(ProcPurchaseOrder).where(ProcPurchaseOrder.po_number == po_number)
    )
    if not po:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase order not found")
    result = await db.execute(select(ProcPoItem).where(ProcPoItem.po_id == po.id))
    items = list(result.scalars().all())
    return po, items


async def list_purchase_orders(
    db: AsyncSession,
    search: str | None = None,
    vendor_code: str | None = None,
    requestor_email: str | None = None,
):
    """Flat list used by export."""
    q = select(ProcPurchaseOrder).order_by(ProcPurchaseOrder.created_at.desc())
    if vendor_code:
        q = q.where(ProcPurchaseOrder.vendor_code == vendor_code)
    if requestor_email:
        q = q.join(ProcIndent, ProcPurchaseOrder.indent_id == ProcIndent.id).where(
            ProcIndent.requestor_email == requestor_email
        )
    if search:
        term = f"%{search}%"
        q = q.where(
            or_(
                ProcPurchaseOrder.po_number.ilike(term),
                ProcPurchaseOrder.status.ilike(term),
                ProcPurchaseOrder.site_id.ilike(term),
            )
        )
    result = await db.execute(q)
    return list(result.scalars().all())


async def list_purchase_orders_paginated(
    db: AsyncSession,
    search: str | None = None,
    page: int = 1,
    limit: int = 10,
    vendor_code: str | None = None,
    requestor_email: str | None = None,
    status: str | None = None,
    state: str | None = None,
) -> tuple[list[dict], dict, list[str]]:
    """Paginated PO list.

    Returns (po_list, pagination, available_states).

    vendor_code: filter + _indentDetailsPayload with PO items + grnDetails + invoiceDetails.
    requestor_email: filter + _indentDetailsPayload with indent items.
    status: filter by PO status (e.g. GRN_SUBMITTED).
    state: filter by site state/region.
    """
    q = (
        select(
            ProcPurchaseOrder,
            ProcIndent.tracking_no.label("indent_tracking_no"),
            Site.location_name.label("site_name"),
            Site.state.label("site_state"),
        )
        .outerjoin(ProcIndent, ProcPurchaseOrder.indent_id == ProcIndent.id)
        .outerjoin(Site, ProcPurchaseOrder.site_id == cast(Site.id, String))
        .order_by(ProcPurchaseOrder.created_at.desc())
    )

    if vendor_code:
        q = q.where(ProcPurchaseOrder.vendor_code == vendor_code)
    if requestor_email:
        q = q.where(ProcIndent.requestor_email == requestor_email)
    if status:
        q = q.where(ProcPurchaseOrder.status == status)
    if state:
        q = q.where(Site.state == state)

    if search:
        term = f"%{search}%"
        q = q.where(
            or_(
                ProcPurchaseOrder.po_number.ilike(term),
                ProcPurchaseOrder.status.ilike(term),
                Site.location_name.ilike(term),
                ProcIndent.tracking_no.ilike(term),
            )
        )

    count_q = select(func.count()).select_from(q.subquery())
    total = await db.scalar(count_q) or 0

    offset = (page - 1) * limit
    result = await db.execute(q.offset(offset).limit(limit))
    rows = list(result.all())

    def _fmt_date(d) -> str | None:
        if d is None:
            return None
        if hasattr(d, "strftime"):
            return d.strftime("%Y-%m-%d")
        return str(d)

    # ── available_states (vendor view only, ignores state filter) ──
    available_states: list[str] = []
    if vendor_code:
        from sqlalchemy import distinct as _distinct
        states_q = (
            select(_distinct(Site.state))
            .join(ProcPurchaseOrder, cast(Site.id, String) == ProcPurchaseOrder.site_id)
            .where(ProcPurchaseOrder.vendor_code == vendor_code)
            .where(Site.state.isnot(None))
        )
        states_result = await db.execute(states_q)
        available_states = sorted([r[0] for r in states_result.all() if r[0]])

    # ── Requestor view: batch-load indent items ──
    indent_items_map: dict[str, list[dict]] = {}
    if requestor_email:
        indent_ids = [po.indent_id for po, *_ in rows if po.indent_id]
        if indent_ids:
            from app.procurement.indents.models import ProcIndentItem
            items_result = await db.execute(
                select(ProcIndentItem).where(ProcIndentItem.indent_id.in_(indent_ids))
            )
            for item in items_result.scalars().all():
                indent_items_map.setdefault(str(item.indent_id), []).append({
                    "productCode": item.product_code,
                    "productName": item.product_name,
                    "quantity": float(item.quantity),
                    "unitPrice": float(item.unit_price),
                    "totalPrice": float(item.total_price),
                })

    # ── Batch-load PO items + indent metadata (all views) ──
    po_items_map: dict[uuid.UUID, list] = {}
    indent_meta_map: dict[str, dict] = {}
    po_ids_list = [po.id for po, *_ in rows]
    if po_ids_list:
        pi_result = await db.execute(
            select(ProcPoItem).where(ProcPoItem.po_id.in_(po_ids_list))
        )
        for item in pi_result.scalars().all():
            po_items_map.setdefault(item.po_id, []).append(item)

    indent_ids_all = [po.indent_id for po, *_ in rows if po.indent_id]
    if indent_ids_all:
        ind_result = await db.execute(
            select(
                ProcIndent.id,
                ProcIndent.is_monthly,
                ProcIndent.for_month,
                ProcIndent.total_value,
            ).where(ProcIndent.id.in_(indent_ids_all))
        )
        for row in ind_result.all():
            indent_meta_map[str(row[0])] = {
                "isMonthly": row[1],
                "forMonth": row[2],
                "totalValue": float(row[3]),
            }

    # ── GRN batch load (all views) ──
    po_numbers_page = [po.po_number for po, *_ in rows]
    grn_map: dict[str, ProcGrn] = {}
    grn_items_map: dict[uuid.UUID, list] = {}
    grn_photos_map: dict[uuid.UUID, list[str]] = {}
    if po_numbers_page:
        grns_result = await db.execute(
            select(ProcGrn).where(ProcGrn.po_number.in_(po_numbers_page))
        )
        grns = list(grns_result.scalars().all())
        for grn in grns:
            grn_map[grn.po_number] = grn
        if grns:
            grn_ids = [grn.id for grn in grns]
            gi_result = await db.execute(
                select(ProcGrnItem).where(ProcGrnItem.grn_id.in_(grn_ids))
            )
            for gi in gi_result.scalars().all():
                grn_items_map.setdefault(gi.grn_id, []).append(gi)
            gp_result = await db.execute(
                select(ProcGrnPhoto).where(ProcGrnPhoto.grn_id.in_(grn_ids))
            )
            for gp in gp_result.scalars().all():
                grn_photos_map.setdefault(gp.grn_id, []).append(gp.photo_url)

    # ── Invoice batch load (all views) ──
    invoice_details_map: dict[str, dict | None] = {}
    if po_numbers_page:
        from app.procurement.invoices.models import ProcInvoice, ProcInvoicePoLink
        lnk_result = await db.execute(
            select(ProcInvoicePoLink).where(
                ProcInvoicePoLink.po_number.in_(po_numbers_page)
            )
        )
        lnks = list(lnk_result.scalars().all())
        if lnks:
            inv_uuids = list({lnk.invoice_id for lnk in lnks})
            invs_result = await db.execute(
                select(ProcInvoice).where(ProcInvoice.id.in_(inv_uuids))
            )
            inv_by_uuid = {inv.id: inv for inv in invs_result.scalars().all()}
            for lnk in lnks:
                inv = inv_by_uuid.get(lnk.invoice_id)
                if inv:
                    invoice_details_map[lnk.po_number] = {
                        "invoiceId": inv.invoice_id,
                        "invoiceNo": inv.invoice_no,
                        "status": inv.status,
                        "billAmount": float(inv.bill_amount),
                        "billUrl": inv.bill_url,
                    }

    po_list = []
    for row in rows:
        po, indent_tracking_no, site_name, site_state = row[0], row[1], row[2], row[3]
        po_dict = {
            "materialRequestId": indent_tracking_no,
            "siteName": site_name or po.site_id,
            "region": site_state,
            "dcNumber": po.dc_number,
            "poNumber": po.po_number,
            "dcDate": _fmt_date(po.dc_date),
            "poDate": _fmt_date(po.po_date),
            "deliveryType": po.delivery_type,
            "tat": po.tat,
            "expectedDeliveryDate": _fmt_date(po.expected_delivery_date),
            "status": po.status,
            "courierName": po.courier_name,
            "podNumber": po.pod_number,
            "dateOfDelivery": _fmt_date(po.date_of_delivery),
            "podImageUrl": po.pod_image_url,
            "signedPodUrl": po.signed_pod_url,
            "signedDcUrl": po.signed_dc_url,
            "signedDcISmartUrl": po.signed_dc_ismart_url,
            "tatStatus": po.tat_status,
            "reason": po.reason,
        }

        # _indentDetailsPayload (all views — requestor gets indent items, others get PO items)
        if po.indent_id:
            if requestor_email:
                po_dict["_indentDetailsPayload"] = {
                    "trackingNo": indent_tracking_no,
                    "items": indent_items_map.get(str(po.indent_id), []),
                }
            else:
                meta = indent_meta_map.get(str(po.indent_id), {})
                po_dict["_indentDetailsPayload"] = {
                    "trackingId": indent_tracking_no,
                    "items": [
                        {
                            "productCode": item.product_code,
                            "productName": item.product_name,
                            "quantity": float(item.quantity),
                            "siteName": site_name or po.site_id,
                            "landedPrice": float(item.landed_price),
                        }
                        for item in po_items_map.get(po.id, [])
                    ],
                    "isMonthly": meta.get("isMonthly", False),
                    "forMonth": meta.get("forMonth", ""),
                    "totalValue": meta.get("totalValue", 0.0),
                }

        # GRN details (all views)
        grn = grn_map.get(po.po_number)
        po_dict["grnDetails"] = None
        if grn:
            po_dict["grnDetails"] = {
                "items": [
                    {
                        "itemId": gi.item_id,
                        "itemName": gi.item_name,
                        "orderedQuantity": float(gi.ordered_quantity),
                        "receivedQuantity": float(gi.received_quantity),
                        "isAccepted": gi.is_accepted,
                    }
                    for gi in grn_items_map.get(grn.id, [])
                ],
                "signedDc": grn.signed_dc_url,
                "comments": grn.comments,
                "packagingImages": grn_photos_map.get(grn.id, []),
            }

        # Invoice details (all views)
        po_dict["invoiceDetails"] = invoice_details_map.get(po.po_number)

        po_list.append(po_dict)

    pagination = {
        "currentPage": page,
        "totalPages": math.ceil(total / limit) if limit else 1,
        "totalItems": total,
    }
    return po_list, pagination, available_states


async def update_purchase_order(
    db: AsyncSession,
    po_number: str,
    data: PoUpdateData,
    pod_image_url: str | None = None,
    signed_pod_url: str | None = None,
    signed_dc_url: str | None = None,
):
    po = await db.scalar(
        select(ProcPurchaseOrder).where(ProcPurchaseOrder.po_number == po_number)
    )
    if not po:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase order not found")

    # --- Validations ---

    # Date cannot be in the future
    if data.dateOfDelivery and data.dateOfDelivery.date() > datetime.now(IST).date():
        raise HTTPException(status_code=400, detail="dateOfDelivery cannot be a future date")

    # "Delivered" status requires all documents
    if data.status == "Delivered":
        errors = []
        if not data.dateOfDelivery:
            errors.append("dateOfDelivery is required when status is 'Delivered'")
        if not data.podNumber:
            errors.append("podNumber is required when status is 'Delivered'")
        if not pod_image_url and not po.pod_image_url:
            errors.append("podImage file is required when status is 'Delivered'")
        if not signed_pod_url and not po.signed_pod_url:
            errors.append("signedPod file is required when status is 'Delivered'")
        if not signed_dc_url and not po.signed_dc_url:
            errors.append("signedDc file is required when status is 'Delivered'")
        if errors:
            raise HTTPException(status_code=400, detail="; ".join(errors))

    # Courier requires courierName
    if data.deliveryType == "Courier" and not data.courierName:
        raise HTTPException(status_code=400, detail="courierName is required when deliveryType is 'Courier'")

    # TAT status calculation + reason requirement
    tat_status = po.tat_status
    if po.expected_delivery_date and data.dateOfDelivery:
        exp_date = po.expected_delivery_date
        if hasattr(exp_date, "date"):
            exp_date = exp_date.date()
        delivery_date = data.dateOfDelivery.date()
        if delivery_date <= exp_date:
            tat_status = "Within TAT"
        else:
            tat_status = "Out of TAT"
            if not data.reason:
                raise HTTPException(
                    status_code=400,
                    detail="reason is required when delivery is out of TAT",
                )

    # --- Apply updates ---
    if data.status is not None:
        po.status = data.status
    if data.deliveryType is not None:
        po.delivery_type = data.deliveryType
    if data.courierName is not None:
        po.courier_name = data.courierName
    if data.podNumber is not None:
        po.pod_number = data.podNumber
    if data.dateOfDelivery is not None:
        po.date_of_delivery = data.dateOfDelivery
    if data.reason is not None:
        po.reason = data.reason
    if tat_status is not None:
        po.tat_status = tat_status
    if pod_image_url:
        po.pod_image_url = pod_image_url
    if signed_pod_url:
        po.signed_pod_url = signed_pod_url
    if signed_dc_url:
        po.signed_dc_url = signed_dc_url

    po.updated_at = datetime.now(IST)
    await db.commit()
    await db.refresh(po)
    return po


async def submit_grn(
    db: AsyncSession,
    po_number: str,
    data: GrnData,
    signed_dc_url: str,
    photo_urls: list[str],
):
    po = await db.scalar(
        select(ProcPurchaseOrder).where(ProcPurchaseOrder.po_number == po_number)
    )
    if not po:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Purchase order not found")

    existing = await db.scalar(select(ProcGrn).where(ProcGrn.po_id == po.id))
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="GRN already submitted for this PO")

    # Validate comments required when predefinedComment is OTHER
    if data.predefinedComment == "OTHER" and not data.comments:
        raise HTTPException(status_code=400, detail="comments is required when predefinedComment is 'OTHER'")

    # Look up PO items to get item names and ordered quantities
    result = await db.execute(select(ProcPoItem).where(ProcPoItem.po_id == po.id))
    po_items = {item.item_id: item for item in result.scalars().all()}

    grn = ProcGrn(
        po_id=po.id,
        po_number=po_number,
        requestor_email=data.requestorEmail,
        predefined_comment=data.predefinedComment,
        comments=data.comments,
        signed_dc_url=signed_dc_url,
    )
    db.add(grn)
    await db.flush()

    for item in data.items:
        po_item = po_items.get(item.itemId)
        item_name = po_item.product_name if po_item else item.itemId
        ordered_qty = float(po_item.quantity) if po_item else 0
        db.add(ProcGrnItem(
            grn_id=grn.id,
            item_id=item.itemId,
            item_name=item_name,
            ordered_quantity=ordered_qty,
            received_quantity=item.receivedQuantity,
            is_accepted=item.isAccepted,
        ))

    for url in photo_urls:
        db.add(ProcGrnPhoto(grn_id=grn.id, photo_url=url))

    # GRN submission changes PO status to GRN_SUBMITTED
    po.status = "GRN_SUBMITTED"
    po.updated_at = datetime.now(IST)

    await db.commit()
    return grn


async def get_grn_with_details(db: AsyncSession, grn_id: uuid.UUID):
    result_items = await db.execute(select(ProcGrnItem).where(ProcGrnItem.grn_id == grn_id))
    items = list(result_items.scalars().all())
    result_photos = await db.execute(select(ProcGrnPhoto).where(ProcGrnPhoto.grn_id == grn_id))
    photos = list(result_photos.scalars().all())
    return items, photos


async def export_purchase_orders(
    db: AsyncSession,
    search: str | None = None,
    vendor_code: str | None = None,
    requestor_email: str | None = None,
):
    from openpyxl import Workbook
    from app.shared.excel_utils import workbook_to_streaming_response

    pos = await list_purchase_orders(
        db, search=search, vendor_code=vendor_code, requestor_email=requestor_email
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = [
        "PO Number", "Site ID", "Status", "Total Value (INR)",
        "PO Date", "Expected Delivery", "Delivery Type",
        "Courier", "POD Number", "Date of Delivery",
    ]
    ws.append(headers)

    from app.shared.excel_utils import HEADER_FONT, HEADER_FILL, HEADER_BORDER
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)

    for po in pos:
        ws.append([
            po.po_number,
            po.site_id,
            po.status,
            float(po.total_value),
            po.po_date.strftime("%Y-%m-%d") if po.po_date else "",
            po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else "",
            po.delivery_type or "",
            po.courier_name or "",
            po.pod_number or "",
            po.date_of_delivery.strftime("%Y-%m-%d") if po.date_of_delivery else "",
        ])

    ws.freeze_panes = "A2"
    from datetime import date as _date
    filename = f"Purchase_Orders_{_date.today().strftime('%Y-%m-%d')}.xlsx"
    return workbook_to_streaming_response(wb, filename)


async def download_po_excel(db: AsyncSession, po_number: str):
    """Generate an Excel file for a single PO with its line items."""
    from openpyxl import Workbook
    from app.shared.excel_utils import workbook_to_streaming_response, HEADER_FONT, HEADER_FILL, HEADER_BORDER

    po, items = await get_po_with_items(db, po_number)
    vendor_name = await get_vendor_name(db, po.vendor_code)

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # PO header info
    info_rows = [
        ("PO Number", po.po_number),
        ("Vendor", vendor_name or po.vendor_code or ""),
        ("Site ID", po.site_id),
        ("PO Date", po.po_date.strftime("%Y-%m-%d") if po.po_date else ""),
        ("Expected Delivery", po.expected_delivery_date.strftime("%Y-%m-%d") if po.expected_delivery_date else ""),
        ("Status", po.status),
        ("Total Value", float(po.total_value)),
    ]
    for row_idx, (label, value) in enumerate(info_rows, start=1):
        ws.cell(row=row_idx, column=1, value=label).font = HEADER_FONT
        ws.cell(row=row_idx, column=2, value=value)

    # Items table starts after a blank row
    items_start = len(info_rows) + 2
    headers = ["Item ID", "Product Code", "Product Name", "Quantity", "Landed Price", "Total Amount"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=items_start, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)

    for i, item in enumerate(items, start=1):
        ws.append([
            item.item_id,
            item.product_code or "",
            item.product_name,
            float(item.quantity),
            float(item.landed_price),
            float(item.total_amount),
        ])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    return workbook_to_streaming_response(wb, f"{po_number}.xlsx")


async def get_vendor_name(db: AsyncSession, vendor_code: str | None) -> str | None:
    if not vendor_code:
        return None
    vendor = await db.get(ProcVendor, vendor_code)
    return vendor.company_name if vendor else None

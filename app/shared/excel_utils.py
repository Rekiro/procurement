"""Shared Excel utilities for template generation, parsing, and streaming."""

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi.responses import Response


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_BORDER = Border(bottom=Side(style="thin"))


def create_template_workbook(
    headers: list[str],
    dropdowns: dict[str, list[str]] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    dropdowns = dropdowns or {}

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)

    hidden_sheet = None
    hidden_col = 0

    for header, values in dropdowns.items():
        if header not in headers:
            continue
        col_idx = headers.index(header) + 1
        col_letter = get_column_letter(col_idx)

        formula_str = ",".join(values)
        if len(formula_str) > 200:
            if hidden_sheet is None:
                hidden_sheet = wb.create_sheet("_lists")
                hidden_sheet.sheet_state = "hidden"
            hidden_col += 1
            hcol_letter = get_column_letter(hidden_col)
            for row_i, val in enumerate(values, start=1):
                hidden_sheet.cell(row=row_i, column=hidden_col, value=val)
            formula = f"_lists!${hcol_letter}$1:${hcol_letter}${len(values)}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        else:
            dv = DataValidation(type="list", formula1=f'"{formula_str}"', allow_blank=True)

        dv.error = f"Please select a valid {header}"
        dv.errorTitle = "Invalid Value"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1000")

    ws.freeze_panes = "A2"
    return wb


def workbook_to_streaming_response(wb: Workbook, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def parse_upload_to_rows(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    if filename.lower().endswith(".csv"):
        return _parse_csv(file_bytes)
    return _parse_xlsx(file_bytes)


def _parse_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header_row = next(rows_iter, None)
    if header_row is None:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    result = []
    for row in rows_iter:
        values = list(row)
        if all(v is None for v in values):
            continue
        record = {}
        for i, header in enumerate(headers):
            if header:
                record[header] = values[i] if i < len(values) else None
        result.append(record)

    wb.close()
    return result


def _parse_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]
